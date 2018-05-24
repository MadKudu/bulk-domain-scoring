import argparse
import re

from openpyxl import load_workbook

from utils import open_xls_as_xlsx


def zip_results(filename: str, score_type: str, column_idx: int):
    regex = re.compile('(?:@)?(?P<tld>[\w\-]+\.\w+)')

    print("Welcome to the results zipper! Wait for the xlsx to load.")
    if re.search('\.xlsx$', filename):
        workbook = load_workbook(filename=filename, keep_vba=False)
        result_filename = filename.replace(".xlsx", ".csv")
    elif re.search('\.xls$', filename):
        workbook = open_xls_as_xlsx(filename)
        result_filename = filename.replace(".xls", ".csv")
    else:
        print("Unsupported file format!")
        exit(1)
    sheet = workbook.active

    print("File loaded. results/{} will be zipped to it.".format(result_filename))
    with open("results/" + result_filename, "rb") as result:
        result.seek(0)
        rows = sheet.max_row
        columns = sheet.max_column
        skip_empty = 2
        for row in sheet['A2:B256']:
            if not row[0].value:
                skip_empty += 1
            else:
                break
        for line in range(skip_empty, rows):
            if line % 100 == 0:
                print("Zipped {0:.3f}%".format(line / (rows * 1.) * 100.))
            result_line = result.readline()
            if not result_line:
                # xlsx had some invalid lines, some where skiped
                # resulting in a discrepancy between the two files lines numbers
                break
            domain, segment, score = result_line.decode("utf-8").split(",")
            mail = sheet['{}{}'.format(column_idx, line)].value
            if not mail or not regex.search(mail):
                result.seek(-len(result_line), 1)  # go back to read the same line again
                continue  # skip invalid entries, like in bulk_score
            if domain not in mail:
                print("ERROR! Zipping on weird data (trying to zip {} with {})".format(mail, domain))
                exit(1)
            sheet.cell(line, columns + 1).value = segment
            sheet.cell(line, columns + 2).value = int(score)
    filename_with_results = filename.replace(".xls", "_with-results.xls")
    print("Now saving to {}, this might take several minutes...".format(filename_with_results))
    workbook.save(filename_with_results)
    print("You're good to go!")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Sends bulk persons to be scored.')
    parser.add_argument("--filename", help="xlsx file containing all the persons to score", required=True)
    parser.add_argument("--score_type", help="which score to use: either by domain or by personal email", required=True, choices=['domain', 'mail'])
    parser.add_argument("--column_idx", help="domain/mail column idx (i.e: BQ)", required=True)

    zip_results(**vars(parser.parse_args()))
