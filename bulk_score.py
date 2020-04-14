import argparse
import asyncio
import json
import logging
import re
import sys

import aiohttp
from asyncio_throttle import Throttler
from openpyxl import load_workbook
from utils import open_xls_as_xlsx

throttler = Throttler(rate_limit=500, period=60)

API_DOMAIN_URL = "https://api.madkudu.com/v1/companies"
API_PERSON_URL = "https://api.madkudu.com/v1/persons"

logger = logging.getLogger('bulk_score')
logger.addHandler(logging.StreamHandler(sys.stdout))
logger.setLevel(logging.DEBUG)


async def get(mode, api_key, param):

    params = {mode: param}
    url = API_DOMAIN_URL if mode == "domain" else API_PERSON_URL
    try:
        auth = aiohttp.BasicAuth(
            login=api_key,
            password=''
        )
        async with aiohttp.ClientSession() as session:
            async with throttler:
                async with session.get(url=url, auth=auth, params=params) as response:
                    return await response.json()
    except Exception as e:
        print(e)
        # print("Unable to get url {} due to {}.".format(url, e.__class__))


async def run_xls(filename: str, api_key: str, score_type: str, column_idx: int):
    print("Welcome to the bulk persons searcher! Wait for the xlsx to load.")
    if re.search(r'\.xlsx$', filename):
        workbook = load_workbook(filename=filename, keep_vba=False)
        result_filename = filename.replace(".xlsx", ".csv")
    elif re.search(r'\.xls$', filename):
        workbook = open_xls_as_xlsx(filename)
        result_filename = filename.replace(".xls", ".csv")
    else:
        print("Unsupported file format!")
        exit(1)

    sheet = workbook.active

    values_to_score = []
    values_scored = {}

    async def write_to_file(values_to_score):
        results = await asyncio.gather(*[get(score_type, api_key, value_to_score) for value_to_score in values_to_score])
        for result in results:
            if result and 'properties' in result:
                customer_fit_result = result['properties']['customer_fit']
                values_scored[result[score_type]] = customer_fit_result
                if 'top_signals_formatted' in customer_fit_result:
                    new_row = "{},{},{},{}\"\n".format(
                        result[score_type],
                        customer_fit_result['segment'],
                        customer_fit_result['score'],
                        customer_fit_result['top_signals_formatted'])
                else:
                    new_row = "{},{},{}\"\n".format(result[score_type], customer_fit_result['segment'], customer_fit_result['score'])
                readcsv.write(new_row)

        readcsv.flush()

    print("File loaded. Results will be saved to results/{}.".format(result_filename))
    with open("results/" + result_filename, "a+") as readcsv:
        readcsv.seek(0)
        full_csv = readcsv.read()
        try:
            rows = sheet.max_row
            for line in range(2, rows):
                email_or_domain = sheet['{}{}'.format(column_idx, line)].value

                if email_or_domain not in full_csv:
                    print("scoring: " + email_or_domain)

                    if email_or_domain not in values_scored:
                        values_to_score.append(email_or_domain)

                    if len(values_to_score) == 100:
                        await write_to_file(values_to_score)
                        values_to_score = []
                        print("Currently at {}%".format(line / (rows * 1.) * 100.))

        except Exception:
            readcsv.flush()
            logger.exception("Exception met. Relaunch to resume!\n")
            exit(1)

        if len(values_to_score) > 0:
            await write_to_file(values_to_score)
        exit(0)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Sends bulk persons to be scored.')
    parser.add_argument("--filename", help="xlsx file containing all the persons to score", required=True)
    parser.add_argument("--api_key", help="api key", required=True)
    parser.add_argument("--score_type", help="which score to use: either by domain or by personal email", required=True, choices=['domain', 'email'])
    parser.add_argument("--column_idx", help="domain/email column idx (i.e: BQ)", required=True)

    loop = asyncio.get_event_loop()
    loop.run_until_complete(run_xls(**vars(parser.parse_args())))
